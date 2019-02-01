import threading
import os
import re
from openpyxl import load_workbook
from DOEValidations.ValidateFiles import Validations
from DOEValidations.OutputValidations import ResultsOutput
from time import gmtime, strftime
import shutil
from tkinter import *
import tkinter.ttk as ttk

# -----------------------------------------------------------------------------------------------
def ProcessFile(data , TestItem, iCounter, version):
    print("---------------------------------------------------------------------------------------------------------------")
    ValkinFile = "Valkin_" + strftime("%Y-%m-%d_%H_%M_", gmtime()) + "_" + str(iCounter) + ".dat"
    batchfilename = ValkinFile.replace(".dat", ".bat")
        #"batchRun_" + strftime("%Y-%m-%d_%H_%M", gmtime()) + "_" + str(iCounter) + ".bat"

    file = open(directory + ValkinFile , "w")
    batchfile = open(directory + batchfilename, "w")
    batchfile.write("cd " + os.path.dirname(os.path.realpath(__file__)) + "\\" + directory + "\n")
    batchfile.write("call valkin -V " + version + " -s " + ValkinFile + "\n")
    iCounter2 = 0
    for x in TestItem:
       # print(FactorsList[iCounter2])
        data = re.sub(VarNames[iCounter2] + '=\"\([0-9,.]+\)\[[A-z]+\]\"',
                      VarNames[iCounter2] + '="(' + str(format(float(x), '.2f')) + ')[' + FactorsList[iCounter2] + ']"',
                      data)
        iCounter2 = iCounter2 + 1
    file.write(data)
    file.close()
    batchfile.close()
    print(str(iCounter) + " files created.")
    os.system(os.path.dirname(os.path.realpath(__file__)) + "\\" + directory + batchfilename)
    print("---------------------------------------------------------------------------------------------------------------")
# -----------------------------------------------------------------------------------------------

strFileName = '../DataFiles/DOE_rampHeight.xlsx'
wb = load_workbook(filename=strFileName, data_only=True, read_only=True)
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('DataSheet')
row_count = sheet.max_row
column_count = sheet.max_column
#print(row_count)
#print(column_count)

DataList = []
strInputFile = sheet['B1'].value
strValkinVersion = sheet['B2'].value
print("Valkin Version - " + strValkinVersion + "\n" + "Input File - " + strInputFile)
FactorsList = []
VarNames = []
for i in range(4, row_count + 1):
    tempList = []
    tempList.append(sheet['B' + str(i)].value)
    VarNames.append(sheet['B' + str(i)].value)
    tempList.append(sheet['C' + str(i)].value)
    tempList.append(sheet['D' + str(i)].value)
    tempList.append(sheet['E' + str(i)].value)
    tempList.append(sheet['F' + str(i)].value)
    FactorsList.append(sheet['F' + str(i)].value)
    DataList.append(tempList)
#print(DataList)
TestData = Validations().GetCombinationMatrix(DataList)
#print(TestData)

wb._archive.close()

# Current Working Directory
dir_src = os.getcwd()
# creating a new directory
directory = "DOE_Models\\" + strftime("%Y-%m-%d_%H_%M", gmtime()) + "\\"
if not os.path.exists(directory):
    os.makedirs(directory)

#------------------------------------------------------------------------------------------
# Copying supporting files to the folder
#------------------------------------------------------------------------------------------

for filename in os.listdir(dir_src):
    if filename.endswith('.prn'):
        shutil.copy(filename, directory)

# Python Threading

iCounter = 0
Processes = []
file = open(strInputFile, 'r')
data = file.read()
for TestItem in TestData:
    iCounter = iCounter + 1
    Processes.append(threading.Thread(target=ProcessFile, args=(data, TestItem, iCounter, strValkinVersion)))

ProcessGroups = {}
iCounter = 0
iGrpCounter = 0
iMaxLimit = 3

for process in Processes :
    if(iCounter%iMaxLimit == 0):
        iGrpCounter = iGrpCounter + 1
        ProcessGroups["Group_" + str(iGrpCounter)] = []
    ProcessGroups["Group_" + str(iGrpCounter)].append(process)
    iCounter = iCounter + 1

for group in ProcessGroups.keys():
    Processes = ProcessGroups[group]
    for process in Processes:
        process.start()

    for process in Processes:
        process.join()
    print(group + " execution complete.")
print("Done")


# --------------------------------------------------------------------------------------------------------------------------------------------------
# Result post-processing
# -------------------------------------------------------------------------------------------------------------------------------------------------

os.chdir(dir_src + "\\" + directory)
# print (os.getcwd())

dir_post = "DOE_Results\\"
if not os.path.exists(dir_post):
    os.makedirs(dir_post)

for rplotfile in os.listdir(os.getcwd()):
    if rplotfile.endswith('.rp') and (not rplotfile.endswith("_zero.rp")):
        shutil.copy(rplotfile, dir_post)

for sdffile in os.listdir(os.getcwd()):
    if sdffile.endswith('.sdf'):
        os.remove(sdffile)

os.chdir(dir_post)

ResultsOutput.ValidateAccelerationData(os.getcwd(), 1.5)
